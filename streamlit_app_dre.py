
import streamlit as st
import pandas as pd
from pathlib import Path

st.set_page_config(page_title="DRE – Elicon", layout="wide")

# -----------------------------
# Utilities
# -----------------------------
PT_MONTHS = ["janeiro","fevereiro","março","abril","maio","junho",
             "julho","agosto","setembro","outubro","novembro","dezembro"]

def month_label(dt: pd.Timestamp) -> str:
    if pd.isna(dt):
        return ""
    m = PT_MONTHS[int(dt.month) - 1]
    return f"{m}/{int(dt.year) % 100:02d}"

@st.cache_data(show_spinner=False)
def load_data(path: str, preferred_sheet: str = "bd"):
    # Abre o arquivo e resolve a aba de forma resiliente
    xls = pd.ExcelFile(path, engine="openpyxl")
    sheetnames = [str(s).strip() for s in xls.sheet_names]
    target = None
    for s in sheetnames:
        if s.lower() == preferred_sheet.lower():
            target = s
            break
    if target is None:
        target = sheetnames[0]
    df = pd.read_excel(xls, sheet_name=target, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]

    # Padroniza MÊS REF
    if "MÊS REF" in df.columns:
        df["MÊS REF"] = pd.to_datetime(df["MÊS REF"], errors="coerce")
        df["PERIODO_LABEL"] = df["MÊS REF"].apply(month_label)
    return df, target

def money(x):
    try:
        return f"R$ {float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"

def perc(x):
    try:
        return (f"{float(x)*100:,.2f}%").replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0,00%"

def pct(dividend, divisor):
    if divisor == 0:
        return 0.0
    return dividend / divisor

def resolve_col(df: pd.DataFrame, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None

def resolve_col_ci(df: pd.DataFrame, targets: list[str], fallback_first: bool = True):
    """
    Resolve o nome de uma coluna de forma case-insensitive,
    comparando contra aliases-alvo. Se não encontrar e fallback_first=True,
    retorna a primeira coluna.
    """
    cols = list(df.columns)
    lowered = {str(c): str(c).strip().lower() for c in cols}
    target_norm = [t.strip().lower() for t in targets]
    for cname, low in lowered.items():
        if low in target_norm:
            return cname
    if fallback_first and cols:
        return cols[0]
    return None

# -----------------------------
# Load
# -----------------------------
data_path = Path("BD.xlsx")
if not data_path.exists():
    st.error("Arquivo 'BD.xlsx' não encontrado no diretório do app. Faça o upload em 'Files' do Streamlit Cloud ou adicione ao repo.")
    st.stop()

df, resolved_sheet = load_data(str(data_path), preferred_sheet="bd")

# -----------------------------
# Mapeamento de colunas (candidatos)
# -----------------------------
C_FAT = ["FAT MÊS $", "FAT MES $", "FAT_MES_$", "FAT_MES", "FAT"]
E_DED = ["DEDUÇÕES LEGAIS", "DEDUCOES LEGAIS", "DEDUCOES", "DEDUÇÕES"]
K_SAL = ["SALÁRIO", "SALARIO"]
L_VT  = ["VALE TRANSPORTE", "VT"]
M_VA  = ["VALE ALIMENTAÇÃO", "VALE ALIMENTACAO", "VA"]
N_VR  = ["VALE REFEIÇÃO", "VALE REFEICAO", "VR"]
O_ASS = ["ASSIDUIDADE"]
S_ENC = ["TOTAL ENCARGOS", "ENCARGOS", "TOTAL_ENCARGOS"]
AA_FT = ["FT"]
AB_FR = ["FREELANCE"]
U_RATEIO = ["RATEIO MP", "RATEIO_MP", "RATEIO"]
MCOL = ["MATERIAL DE CONSUMO", "MATERIAL_CONSUMO", "MAT CONSUMO"]

col_fat = resolve_col(df, C_FAT)
col_ded = resolve_col(df, E_DED)
col_sal = resolve_col(df, K_SAL)
col_vt  = resolve_col(df, L_VT)
col_va  = resolve_col(df, M_VA)
col_vr  = resolve_col(df, N_VR)
col_ass = resolve_col(df, O_ASS)
col_enc = resolve_col(df, S_ENC)
col_ft  = resolve_col(df, AA_FT)
col_frl = resolve_col(df, AB_FR)
col_rat = resolve_col(df, U_RATEIO)
col_mco = resolve_col(df, MCOL)

cost_cols = [c for c in [col_sal, col_vt, col_va, col_vr, col_ass, col_enc, col_ft, col_frl, col_rat, col_mco] if c is not None]

# -----------------------------
# Sidebar (Filtros)
# -----------------------------
st.sidebar.header("Parâmetros")
st.sidebar.caption(f"Aba carregada: **{resolved_sheet}**")

col_empresa = "EMPRESA" if "EMPRESA" in df.columns else list(df.columns)[0]

# PRIORIDADE: nova coluna "MÊS REF"; fallback para "TIMES"
if "MÊS REF" in df.columns:
    col_periodo = "MÊS REF"
    periodos_unique = df[col_periodo].dropna().drop_duplicates().sort_values()
    labels = [month_label(pd.to_datetime(x)) for x in periodos_unique]
    idx_default = len(labels) - 1 if len(labels) > 0 else 0
    label_sel = st.sidebar.selectbox("Período (MÊS REF – fim do mês)", labels, index=idx_default)
    periodo_sel_dt = periodos_unique.iloc[labels.index(label_sel)] if len(labels) > 0 else None
else:
    # Fallback legacy
    col_periodo = "TIMES" if "TIMES" in df.columns else ( [c for c in df.columns if c.lower().startswith("time")] + [list(df.columns)[-1]] )[0]
    periodos = sorted(df[col_periodo].dropna().astype(str).unique().tolist())
    label_sel = st.sidebar.selectbox("Período (TIMES)", periodos, index=len(periodos)-1)
    periodo_sel_dt = None  # não usado no fallback

empresas = sorted(df[col_empresa].dropna().astype(str).unique().tolist())
cliente_sel = st.sidebar.selectbox("Cliente", empresas, index=0)

aba = st.sidebar.radio("Visão", ["DRE por Cliente", "DRE Consolidado", "Dashboard", "Orçado x Realizado"], index=0)

with st.sidebar.expander("Dicionário de Dados", expanded=False):
    st.markdown(
        "- EMPRESA (col. A): filtro de cliente  \n"
        "- MÊS REF (col. Y, fim do mês): filtro principal de período  \n"
        "- FAT MÊS $ (col. C) -> Faturamento Bruto  \n"
        "- DEDUÇÕES LEGAIS (col. E) -> Deduções  \n"
        "- SALÁRIO (K), VT (L), VA (M), VR (N), ASSIDUIDADE (O)  \n"
        "- TOTAL ENCARGOS (S), FT (AA), FREELANCE (AB), RATEIO MP (U)  \n"
        "- MATERIAL DE CONSUMO (nome exato)"
    )

# -----------------------------
# Helpers de cálculo
# -----------------------------
def compute_totals(dff: pd.DataFrame):
    fat_bruto = dff[col_fat].fillna(0).sum() if col_fat else 0
    deducoes = dff[col_ded].fillna(0).sum() if col_ded else 0
    fat_liq = fat_bruto - deducoes
    csp = dff[cost_cols].fillna(0).sum().sum() if cost_cols else 0
    mc = fat_liq - csp
    return fat_bruto, deducoes, fat_liq, csp, mc

@st.cache_data(show_spinner=False)
def load_budget(path: str, preferred_sheet: str = "bd"):
    xls = pd.ExcelFile(path, engine="openpyxl")
    sheetnames = [str(s).strip() for s in xls.sheet_names]
    target = None
    for s in sheetnames:
        if s.lower() == preferred_sheet.lower():
            target = s
            break
    if target is None:
        target = sheetnames[0]
    dfb = pd.read_excel(xls, sheet_name=target, engine="openpyxl")
    dfb.columns = [str(c).strip() for c in dfb.columns]
    if "MÊS REF" in dfb.columns:
        dfb["MÊS REF"] = pd.to_datetime(dfb["MÊS REF"], errors="coerce")
        dfb["PERIODO_LABEL"] = dfb["MÊS REF"].apply(month_label)
    return dfb, target

def compute_block(df_block: pd.DataFrame, col_fat, col_ded, cost_cols):
    fat = df_block[col_fat].fillna(0).sum() if col_fat else 0
    ded = df_block[col_ded].fillna(0).sum() if col_ded else 0
    fat_liq = fat - ded
    csp = df_block[cost_cols].fillna(0).sum().sum() if cost_cols else 0
    mc = fat_liq - csp
    return {"FAT BRUTO": fat, "DEDUÇÕES": ded, "FAT LÍQ": fat_liq, "CSP": csp, "MC": mc}

def block_dre(title: str, dff: pd.DataFrame):
    fat_bruto, deducoes, fat_liq, csp, mc = compute_totals(dff)
    if title:
        st.markdown(f"### {title}")
    def linha(label, valor, base_pct, highlight=False):
        c1, c2, c3 = st.columns([2.8, 1.2, 0.8])
        with c1:
            st.markdown(f"**{label}**" if highlight else label)
        with c2:
            st.markdown(f"<div style='text-align:right'>{money(valor)}</div>", unsafe_allow_html=True)
        with c3:
            st.markdown(f"<div style='text-align:right'>{perc(pct(valor, base_pct))}</div>", unsafe_allow_html=True)

    st.divider()
    st.markdown("#### REALIZADO | AV%")
    linha("(+) FATURAMENTO BRUTO", fat_bruto, fat_bruto, highlight=True)
    linha("(–) DEDUÇÕES LEGAIS", deducoes, fat_bruto)
    linha("(=) FATURAMENTO LÍQUIDO", fat_liq, fat_bruto, highlight=True)
    st.divider()
    linha("(–) CSP", csp, fat_bruto, highlight=True)
    if cost_cols:
        for cname in cost_cols:
            linha(f"(–) {cname}", dff[cname].fillna(0).sum(), fat_bruto)
    st.divider()
    linha("(=) MARGEM DE CONTRIBUIÇÃO", mc, fat_bruto, highlight=True)

# -----------------------------
# Layout – Cabeçalho
# -----------------------------
st.title("Elicon – DRE (Streamlit)")
st.caption("Leitura da base 'BD.xlsx' com aba autodetectada e período por 'MÊS REF' (fim do mês).")

# -----------------------------
# Abas
# -----------------------------
if aba in ["DRE por Cliente", "DRE Consolidado"]:
    # Filtragem base
    if aba == "DRE por Cliente":
        if "MÊS REF" in df.columns:
            dff = df[(df[col_empresa].astype(str) == str(cliente_sel)) & (df["MÊS REF"] == periodo_sel_dt)].copy()
            titulo = f"DRE – {cliente_sel} | {label_sel}"
        else:
            dff = df[(df[col_empresa].astype(str) == str(cliente_sel)) & (df[col_periodo].astype(str) == str(label_sel))].copy()
            titulo = f"DRE – {cliente_sel} | {label_sel}"
    else:
        if "MÊS REF" in df.columns:
            dff = df[df["MÊS REF"] == periodo_sel_dt].copy()
            titulo = f"DRE – Consolidado | {label_sel}"
        else:
            dff = df[(df[col_periodo].astype(str) == str(label_sel))].copy()
            titulo = f"DRE – Consolidado | {label_sel}"

    st.subheader(titulo)
    block_dre("", dff)

    with st.expander("Ver base filtrada (controle/QA)"):
        st.dataframe(dff, use_container_width=True)

    st.caption("Origem dos dados: colunas sinalizadas no template. Percentuais = valor ÷ FATURAMENTO BRUTO.")

elif aba == "Dashboard":
    # -----------------------------
    # DASHBOARD
    # -----------------------------
    st.subheader(f"Dashboard – {label_sel if 'label_sel' in locals() else ''}")

    # Toggle para Margem: % ou R$, e Top-N
    c1, c2 = st.columns([1,1])
    with c1:
        mc_mode = st.radio("Modo de Margem de Contribuição nos Rankings", ["Percentual (%)", "Valor (R$)"], index=0, horizontal=True)
    with c2:
        top_n = st.slider("Top-N", min_value=5, max_value=20, value=10, step=1)

    # Filtra dataframe para o mês corrente selecionado
    if "MÊS REF" in df.columns:
        dfm = df[df["MÊS REF"] == periodo_sel_dt].copy()
    else:
        dfm = df[df[col_periodo].astype(str) == str(label_sel)].copy()

    # Agrupa por EMPRESA (período atual)
    agg_dict = {}
    if col_fat: agg_dict[col_fat] = "sum"
    if col_ded: agg_dict[col_ded] = "sum"
    for c in cost_cols:
        agg_dict[c] = "sum"

    if not agg_dict:
        st.warning("Não foi possível identificar as colunas necessárias para o dashboard.")
    else:
        by_emp = dfm.groupby(col_empresa, dropna=True).agg(agg_dict).fillna(0)
        # Derivados
        by_emp["FATURAMENTO_BRUTO"] = by_emp[col_fat] if col_fat in by_emp.columns else 0
        by_emp["FATURAMENTO_LIQ"] = (by_emp[col_fat] if col_fat in by_emp.columns else 0) - (by_emp[col_ded] if col_ded in by_emp.columns else 0)
        by_emp["CSP"] = by_emp[cost_cols].sum(axis=1) if cost_cols else 0
        by_emp["MARGEM_CONTRIB"] = by_emp["FATURAMENTO_LIQ"] - by_emp["CSP"]
        # Percentual de margem de contribuição sobre Faturamento Bruto
        by_emp["MC_PCT_BRUTO"] = 0.0
        if col_fat and col_fat in by_emp.columns:
            denom = by_emp["FATURAMENTO_BRUTO"].replace(0, pd.NA)
            by_emp["MC_PCT_BRUTO"] = (by_emp["MARGEM_CONTRIB"] / denom).fillna(0.0)

        # Top Faturamento (bruto)
        if col_fat:
            top_fat = by_emp.sort_values(col_fat, ascending=False).head(top_n)
            st.markdown(f"### Top {top_n} – Faturamento Bruto (mês selecionado)")
            st.bar_chart(top_fat[col_fat])
            df_show = top_fat[[col_fat]].rename(columns={col_fat: "FATURAMENTO BRUTO"}).copy()
            df_show["FATURAMENTO BRUTO"] = df_show["FATURAMENTO BRUTO"].map(money)
            st.dataframe(df_show)

        # Top CSP
        st.markdown(f"### Top {top_n} – CSP (mês selecionado)")
        top_csp = by_emp.sort_values("CSP", ascending=False).head(top_n)
        st.bar_chart(top_csp["CSP"])
        df_show = top_csp[["CSP"]].copy()
        df_show["CSP"] = df_show["CSP"].map(money)
        st.dataframe(df_show)

        # Top/Bottom Margens – modo dinâmico
        if mc_mode == "Percentual (%)":
            st.markdown(f"### Top {top_n} – Melhores Margens de Contribuição (%) (mês selecionado)")
            top_mc_best = by_emp.sort_values("MC_PCT_BRUTO", ascending=False).head(top_n)
            st.bar_chart(top_mc_best["MC_PCT_BRUTO"])
            df_best = top_mc_best[["MC_PCT_BRUTO", "MARGEM_CONTRIB", "FATURAMENTO_BRUTO"]].rename(columns={"MC_PCT_BRUTO":"MC % (sobre FAT BRUTO)"}).copy()
            df_best["MC % (sobre FAT BRUTO)"] = df_best["MC % (sobre FAT BRUTO)"].map(perc)
            df_best["MARGEM_CONTRIB"] = df_best["MARGEM_CONTRIB"].map(money)
            df_best["FATURAMENTO_BRUTO"] = df_best["FATURAMENTO_BRUTO"].map(money)
            st.dataframe(df_best)

            st.markdown(f"### Top {top_n} – Piores Margens de Contribuição (%) (mês selecionado)")
            top_mc_worst = by_emp.sort_values("MC_PCT_BRUTO", ascending=True).head(top_n)
            st.bar_chart(top_mc_worst["MC_PCT_BRUTO"])
            df_worst = top_mc_worst[["MC_PCT_BRUTO", "MARGEM_CONTRIB", "FATURAMENTO_BRUTO"]].rename(columns={"MC_PCT_BRUTO":"MC % (sobre FAT BRUTO)"}).copy()
            df_worst["MC % (sobre FAT BRUTO)"] = df_worst["MC % (sobre FAT BRUTO)"].map(perc)
            df_worst["MARGEM_CONTRIB"] = df_worst["MARGEM_CONTRIB"].map(money)
            df_worst["FATURAMENTO_BRUTO"] = df_worst["FATURAMENTO_BRUTO"].map(money)
            st.dataframe(df_worst)
        else:
            st.markdown(f"### Top {top_n} – Maiores Margens de Contribuição (R$) (mês selecionado)")
            top_mc_best = by_emp.sort_values("MARGEM_CONTRIB", ascending=False).head(top_n)
            st.bar_chart(top_mc_best["MARGEM_CONTRIB"])
            df_best = top_mc_best[["MARGEM_CONTRIB", "FATURAMENTO_BRUTO"]].copy()
            df_best["MARGEM_CONTRIB"] = df_best["MARGEM_CONTRIB"].map(money)
            df_best["FATURAMENTO_BRUTO"] = df_best["FATURAMENTO_BRUTO"].map(money)
            st.dataframe(df_best)

            st.markdown(f"### Top {top_n} – Menores Margens de Contribuição (R$) (mês selecionado)")
            top_mc_worst = by_emp.sort_values("MARGEM_CONTRIB", ascending=True).head(top_n)
            st.bar_chart(top_mc_worst["MARGEM_CONTRIB"])
            df_worst = top_mc_worst[["MARGEM_CONTRIB", "FATURAMENTO_BRUTO"]].copy()
            df_worst["MARGEM_CONTRIB"] = df_worst["MARGEM_CONTRIB"].map(money)
            df_worst["FATURAMENTO_BRUTO"] = df_worst["FATURAMENTO_BRUTO"].map(money)
            st.dataframe(df_worst)

        st.divider()
        # Drill-down inline
        st.markdown("## Drill-down de Cliente (mês selecionado)")
        clientes_rank = by_emp.index.tolist()
        if clientes_rank:
            cliente_pick = st.selectbox("Selecione um cliente para ver a DRE do período", clientes_rank, index=0)
            if "MÊS REF" in df.columns:
                dff_drill = df[(df[col_empresa].astype(str) == str(cliente_pick)) & (df["MÊS REF"] == periodo_sel_dt)].copy()
            else:
                dff_drill = df[(df[col_empresa].astype(str) == str(cliente_pick)) & (df[col_periodo].astype(str) == str(label_sel))].copy()
            block_dre(f"DRE – {cliente_pick} | {label_sel}", dff_drill)
        else:
            st.info("Nenhum cliente encontrado no período selecionado.")


elif aba == "Orçado x Realizado":
    st.subheader(f"Orçado x Realizado – {cliente_sel} | {label_sel if 'label_sel' in locals() else ''}")

    # === Fonte ORÇADO: BD CONT NOVO.xlsx / aba BD CONT ===
    budget_path = Path("BD CONT NOVO.xlsx")
    if not budget_path.exists():
        st.error("Arquivo de orçamento 'BD CONT NOVO.xlsx' não encontrado na raiz. Suba o arquivo e recarregue.")
        st.stop()

    # Leitura robusta da aba BD CONT
    import openpyxl
    xls_b = pd.ExcelFile(str(budget_path), engine="openpyxl")
    sheet_bud = None
    for s in xls_b.sheet_names:
        if str(s).strip().upper().startswith("BD CONT"):
            sheet_bud = s
            break
    if sheet_bud is None:
        sheet_bud = xls_b.sheet_names[0]
    dfb = pd.read_excel(xls_b, sheet_name=sheet_bud, engine="openpyxl")
    dfb.columns = [str(c).strip() for c in dfb.columns]

    # Mapas de colunas – ORÇADO (conforme títulos fornecidos)
    bud_cliente = "Cliente" if "Cliente" in dfb.columns else resolve_col_ci(dfb, ["cliente","empresa","contrato"])
    bud_ded = "(-) DEDUÇÕES LEGAIS" if "(-) DEDUÇÕES LEGAIS" in dfb.columns else resolve_col_ci(dfb, ["(-) deduções legais","deduções legais","deducoes legais"])
    bud_sal = "(-) SALÁRIO" if "(-) SALÁRIO" in dfb.columns else resolve_col_ci(dfb, ["(-) salário","salário","salario"])
    bud_vt  = "(-) VALE TRANSPORTE " if "(-) VALE TRANSPORTE " in dfb.columns else resolve_col_ci(dfb, ["(-) vale transporte","vale transporte"])
    bud_va  = "(-) VALE ALIMENTAÇÃO" if "(-) VALE ALIMENTAÇÃO" in dfb.columns else resolve_col_ci(dfb, ["(-) vale alimentação","vale alimentacao","vale alimentação"])
    bud_vr  = "(-) VALE REFEIÇÃO" if "(-) VALE REFEIÇÃO" in dfb.columns else resolve_col_ci(dfb, ["(-) vale refeição","vale refeicao","vale refeição"])
    bud_ass = "(-) ASSIDUIDADE" if "(-) ASSIDUIDADE" in dfb.columns else resolve_col_ci(dfb, ["(-) assiduidade","assiduidade"])
    bud_fat = "(+) FATURAMENTO BRUTO" if "(+) FATURAMENTO BRUTO" in dfb.columns else resolve_col_ci(dfb, ["(+) faturamento bruto","faturamento bruto"])
    bud_mco = "(-) MATERIAL DE CONSUMO" if "(-) MATERIAL DE CONSUMO" in dfb.columns else resolve_col_ci(dfb, ["(-) material de consumo","material de consumo"])
    bud_enc = "(-) TOTAL ENCARGOS" if "(-) TOTAL ENCARGOS" in dfb.columns else resolve_col_ci(dfb, ["(-) total encargos","total encargos","encargos"])

    bud_cost_cols = [c for c in [bud_sal, bud_vt, bud_va, bud_vr, bud_ass, bud_mco, bud_enc] if c]

    # Filtragem: realizado (BD.xlsx) por cliente + período; orçado (BD CONT NOVO.xlsx) só por cliente
    if "MÊS REF" in df.columns:
        dff_real = df[(df[col_empresa].astype(str) == str(cliente_sel)) & (df["MÊS REF"] == periodo_sel_dt)].copy()
    else:
        dff_real = df[(df[col_empresa].astype(str) == str(cliente_sel)) & (df[col_periodo].astype(str) == str(label_sel))].copy()

    dff_bud = dfb[dfb[bud_cliente].astype(str) == str(cliente_sel)].copy()

    # Resolver colunas REALIZADO no BD.xlsx (podem ter nomes levemente distintos)
    real_ded = resolve_col(df, ["DEDUÇÕES LEGAIS","(-) DEDUÇÕES LEGAIS"])
    real_sal = resolve_col(df, ["SALÁRIO","(-) SALÁRIO"])
    real_vt  = resolve_col(df, ["VALE TRANSPORTE ","(-) VALE TRANSPORTE "])
    real_va  = resolve_col(df, ["VALE ALIMENTAÇÃO","(-) VALE ALIMENTAÇÃO"])
    real_vr  = resolve_col(df, ["VALE REFEIÇÃO","(-) VALE REFEIÇÃO"])
    real_ass = resolve_col(df, ["ASSIDUIDADE","(-) ASSIDUIDADE"])
    real_fat = resolve_col(df, ["FAT MÊS $","(+) FATURAMENTO BRUTO"])
    real_mco = resolve_col(df, ["MATERIAL DE CONSUMO","(-) MATERIAL DE CONSUMO"])
    real_enc = resolve_col(df, ["TOTAL ENCARGOS","(-) TOTAL ENCARGOS"])
    real_cost_cols = [c for c in [real_sal, real_vt, real_va, real_vr, real_ass, real_mco, real_enc] if c]

    def sum_col(df_in, col):
        return df_in[col].fillna(0).sum() if col and col in df_in.columns else 0.0

    # Orçado (R$)
    orc_fat = sum_col(dff_bud, bud_fat)
    orc_ded = sum_col(dff_bud, bud_ded)
    orc_fat_liq = orc_fat - orc_ded
    orc_csp = sum([sum_col(dff_bud, c) for c in bud_cost_cols])
    orc_mc = orc_fat_liq - orc_csp

    # Realizado (R$)
    real_fat_v = sum_col(dff_real, real_fat)
    real_ded_v = sum_col(dff_real, real_ded)
    real_fat_liq = real_fat_v - real_ded_v
    real_csp = sum([sum_col(dff_real, c) for c in real_cost_cols])
    real_mc = real_fat_liq - real_csp

    import numpy as np
    def safe_div(a, b):
        return (a / b) if b not in (0, None) else 0.0

    rows = ["FAT BRUTO","DEDUÇÕES","FAT LÍQ","CSP","MC"]
    comp = pd.DataFrame({
        "Linha": rows,
        "Orçado (R$)": [orc_fat, orc_ded, orc_fat_liq, orc_csp, orc_mc],
        "Realizado (R$)": [real_fat_v, real_ded_v, real_fat_liq, real_csp, real_mc],
    }).set_index("Linha")
    comp["Δ (R$)"] = comp["Realizado (R$)"] - comp["Orçado (R$)"]
    comp["Δ (%)"] = comp["Δ (R$)"] / comp["Orçado (R$)"].replace({0: pd.NA})

    comp_show = comp.copy()
    comp_show["Orçado (R$)"] = comp_show["Orçado (R$)"].map(money)
    comp_show["Realizado (R$)"] = comp_show["Realizado (R$)"].map(money)
    comp_show["Δ (R$)"] = comp_show["Δ (R$)"].map(money)
    comp_show["Δ (%)"] = comp_show["Δ (%)"].map(perc)

    c1, c2 = st.columns([1,1])
    with c1:
        st.markdown("### Orçado x Realizado – Tabela Comparativa")
        st.dataframe(comp_show, use_container_width=True)
    with c2:
        st.markdown("### Visual Comparativo por Linha (R$)")
        st.bar_chart(comp[["Orçado (R$)","Realizado (R$)"]])

    st.divider()
    with st.expander("Bases filtradas (QA)"):
        st.markdown("**Realizado (DF base – BD.xlsx)**")
        st.dataframe(dff_real, use_container_width=True)
        st.markdown("**Orçado (DF orçamento – BD CONT NOVO.xlsx)**")
        st.dataframe(dff_bud, use_container_width=True)
